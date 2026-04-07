from typing import List, Dict, Any
from openpyxl import load_workbook

from mk_chat_core import normalize_phone, normalize_state, normalize_birthday

import re
import calendar
import datetime

def _normalize_import_birthday(raw: str) -> str:
    """
    Import-safe birthday parser.

    Returns:
      - YYYY-MM-DD if year is present
      - MM-DD if year is missing
      - "" if invalid

    Examples:
      "August 26" -> "08-26"
      "8/26" -> "08-26"
      "08/26/82" -> "1982-08-26"
      "1982-08-26" -> "1982-08-26"
    """
    s = (raw or "").strip()
    if not s:
        return ""

    # Full ISO date
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        try:
            y, mo, d = map(int, s.split("-"))
            datetime.date(y, mo, d)
            return f"{y:04d}-{mo:02d}-{d:02d}"
        except Exception:
            return ""

    # Numeric formats: MM/DD, MM/DD/YY, MM/DD/YYYY
    m = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?", s)
    if m:
        mo = int(m.group(1))
        d = int(m.group(2))
        y_raw = m.group(3)

        try:
            if y_raw is None:
                datetime.date(2000, mo, d)  # validate month/day only
                return f"{mo:02d}-{d:02d}"
            else:
                y_i = int(y_raw)
                if len(y_raw) == 2:
                    y = 2000 + y_i if y_i <= 29 else 1900 + y_i
                else:
                    y = y_i
                datetime.date(y, mo, d)
                return f"{y:04d}-{mo:02d}-{d:02d}"
        except Exception:
            return ""

    # Month-name formats: "August 26" or "August 26 1982"
    s2 = re.sub(r"[.,]", " ", s)
    s2 = re.sub(r"\s+", " ", s2).strip()

    month_map = {name.lower(): i for i, name in enumerate(calendar.month_name) if name}
    month_map.update({name.lower(): i for i, name in enumerate(calendar.month_abbr) if name})

    parts = s2.split(" ")

    if len(parts) >= 2 and parts[0].lower() in month_map:
        mo = month_map[parts[0].lower()]
        try:
            d = int(parts[1])
        except Exception:
            return ""

        try:
            if len(parts) >= 3:
                year_token = parts[2]
                y_i = int(year_token)
                if len(year_token) == 2:
                    y = 2000 + y_i if y_i <= 29 else 1900 + y_i
                else:
                    y = y_i
                datetime.date(y, mo, d)
                return f"{y:04d}-{mo:02d}-{d:02d}"
            else:
                datetime.date(2000, mo, d)  # validate month/day only
                return f"{mo:02d}-{d:02d}"
        except Exception:
            return ""

    return ""

def _missing_order_fields(customer: Dict[str, Any]) -> list[str]:
    missing = []

    if not (customer.get("first_name") or "").strip():
        missing.append("first_name")
    if not (customer.get("last_name") or "").strip():
        missing.append("last_name")
    if not (customer.get("street") or "").strip():
        missing.append("street")
    if not (customer.get("city") or "").strip():
        missing.append("city")
    if not (customer.get("state") or "").strip():
        missing.append("state")
    if not (customer.get("postal_code") or "").strip():
        missing.append("postal_code")

    return missing

def _clean_str(v) -> str:
    return (str(v).strip() if v is not None else "")


def parse_customer_export_xlsx(path: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = [_clean_str(cell.value) for cell in ws[1]]
    header_map = {h: i for i, h in enumerate(headers)}

    def get(row, name: str) -> str:
        idx = header_map.get(name)
        if idx is None:
            return ""
        return _clean_str(row[idx])

    customers: List[Dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        first = get(row, "First Name")
        last = get(row, "Last Name")

        if not first and not last:
            continue

        addr1 = get(row, "Address Line 1")
        addr2 = get(row, "Address Line 2")

        customer = {
            "first_name": first,
            "last_name": last,
            "birthday": _normalize_import_birthday(get(row, "Birthday")),
            "phone": normalize_phone(get(row, "Phone")),
            "email": get(row, "Email").lower(),
            "street": addr1,
            "street2": addr2,
            "city": get(row, "City"),
            "state": normalize_state(get(row, "State/Territory")),
            "postal_code": get(row, "Postal Code"),
            "country": get(row, "Country"),
        }

        missing = _missing_order_fields(customer)
        customer["is_order_ready"] = len(missing) == 0
        customer["missing_order_fields"] = missing

        customers.append(customer)

    return customers
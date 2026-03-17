from openpyxl import load_workbook

def inspect_customer_file(path: str):
    wb = load_workbook(path)
    ws = wb.active

    # Get headers (first row)
    headers = [cell.value for cell in ws[1]]
    print("\nHEADERS:")
    for i, h in enumerate(headers):
        print(f"{i}: {h}")

    print("\nFIRST 5 ROWS:")
    for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        print(row)


if __name__ == "__main__":
    inspect_customer_file("/tmp/customer_import_test.xlsx")